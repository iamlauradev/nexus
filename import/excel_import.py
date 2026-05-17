"""
Importa el Excel "Doramas, manga y anime.xlsx" a la base de datos del tracker.

Uso:
    python3 excel_import.py [--excel /ruta/al/archivo.xlsx] [--user laura]
    python3 excel_import.py --dry-run   # solo muestra lo que haría
"""
import os
import sys
import re
import json
import logging
import argparse
from pathlib import Path

import openpyxl
import psycopg2
import psycopg2.extras

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

# ── DB config ──────────────────────────────────────────────────────────────────
def _load_env():
    env = Path(__file__).parent.parent / ".env"
    if env.exists():
        for line in env.read_text().splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, _, v = line.partition("=")
                os.environ.setdefault(k.strip(), v.strip())
_load_env()

DB_HOST     = os.environ.get("DB_HOST", "localhost")
DB_PORT     = int(os.environ.get("DB_PORT", 5432))
DB_NAME     = os.environ.get("DB_NAME", "nexusdb")
DB_USER     = os.environ.get("DB_USER", "nexus")
DB_PASSWORD = os.environ.get("DB_PASSWORD", "nexus2024")

# ── Color → rating mapping ─────────────────────────────────────────────────────
# Derived from Excel cell fill colors (col C = user's rating)
COLOR_TO_RATING = {
    'FF674EA7': 'must',          # Purple
    'FF45818E': 'me_encanta',    # Teal
    'FFF1C232': 'muy_bonita',    # Yellow-gold
    'FF6AA84F': 'bonita',        # Green
    'FFE69138': 'bonita',        # Orange (maps to same as green)
    'FFDD7E6B': 'pasable',       # Salmon
    'FFCC4125': 'no_me_gusto',   # Red
    'FF666666': 'abandonado',    # Gray
    'FFFFFFFF': 'sin_valorar',   # White / no fill
    '00000000': 'sin_valorar',   # Transparent
}

def _get_rating(cell) -> str:
    try:
        fill = cell.fill
        if fill and fill.fill_type not in (None, 'none'):
            rgb = fill.fgColor.rgb if hasattr(fill.fgColor, 'rgb') else str(fill.fgColor.value)
            return COLOR_TO_RATING.get(rgb, 'sin_valorar')
    except Exception:
        pass
    return 'sin_valorar'

# ── Helpers ────────────────────────────────────────────────────────────────────
def _clean(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None

def _clean_list(value, sep='\n') -> list[str]:
    if not value:
        return []
    return [x.strip() for x in str(value).split(sep) if x.strip() and not x.strip().startswith('+')]

def _parse_title_year(raw: str | None) -> tuple[str | None, int | None]:
    if not raw:
        return None, None
    text = str(raw).strip()
    m = re.search(r'\((\d{4})\)', text)
    year = int(m.group(1)) if m else None
    title = re.sub(r'\s*\(\d{4}\)\s*', '', text).strip('\n').strip()
    return title or None, year

def _parse_country_network(raw: str | None) -> tuple[str | None, str | None]:
    if not raw:
        return None, None
    parts = [p.strip() for p in str(raw).split('\n') if p.strip()]
    country = parts[0] if parts else None
    # Network lines follow after empty lines
    networks = [p for p in parts[1:] if p]
    network = ', '.join(networks) if networks else None
    return country, network

def _row_to_none(row_vals: list) -> bool:
    return all(v is None or str(v).strip() == '' for v in row_vals[:12])

def _conn():
    return psycopg2.connect(
        host=DB_HOST, port=DB_PORT, dbname=DB_NAME,
        user=DB_USER, password=DB_PASSWORD,
        cursor_factory=psycopg2.extras.RealDictCursor,
    )

# ── DB ops ─────────────────────────────────────────────────────────────────────
def get_or_create_user(cur, username: str) -> int:
    cur.execute("SELECT id FROM users WHERE username = %s", (username,))
    row = cur.fetchone()
    if row:
        return row['id']
    # Create with a placeholder password
    import hashlib, secrets
    salt = secrets.token_hex(16)
    pw_hash = f"{salt}:{hashlib.sha256(f'{salt}changeme'.encode()).hexdigest()}"
    cur.execute(
        "INSERT INTO users (username, display_name, password_hash) VALUES (%s, %s, %s) RETURNING id",
        (username, username.capitalize(), pw_hash)
    )
    uid = cur.fetchone()['id']
    log.info(f"  Usuario creado: {username} (id={uid}) — password inicial: 'changeme'")
    return uid


def upsert_media(cur, data: dict) -> int:
    cur.execute("""
        INSERT INTO media (type, title, title_original, year, genres, synopsis,
            cover_url, duration, country, network, cast_text, external_score, platform)
        VALUES (%(type)s, %(title)s, %(title_original)s, %(year)s, %(genres)s, %(synopsis)s,
            %(cover_url)s, %(duration)s, %(country)s, %(network)s, %(cast_text)s,
            %(external_score)s, %(platform)s)
        ON CONFLICT DO NOTHING
        RETURNING id
    """, data)
    row = cur.fetchone()
    if row:
        return row['id']
    # Already exists — match by title+type
    cur.execute("SELECT id FROM media WHERE title = %(title)s AND type = %(type)s", data)
    return cur.fetchone()['id']


def upsert_entry(cur, user_id: int, media_id: int, data: dict) -> None:
    cur.execute("""
        INSERT INTO user_entries
            (user_id, media_id, status, progress, score, rating_label, notes, platform)
        VALUES (%(uid)s, %(mid)s, %(status)s, %(progress)s, %(score)s,
                %(rating_label)s, %(notes)s, %(platform)s)
        ON CONFLICT (user_id, media_id) DO UPDATE SET
            status=EXCLUDED.status, progress=EXCLUDED.progress, score=EXCLUDED.score,
            rating_label=EXCLUDED.rating_label, notes=EXCLUDED.notes, platform=EXCLUDED.platform,
            updated_at=NOW()
    """, {'uid': user_id, 'mid': media_id, **data})


# ── Sheet parsers ──────────────────────────────────────────────────────────────
def parse_doramas(ws, user_id: int, status: str, cur, dry_run: bool) -> int:
    """Doramas & Doramas WatchList — col layout:
    B=score  C=rating_color  E=name_year  F=duration  G=genres
    H=country_network  I=cast  J=synopsis  K=notes  L=platform
    """
    count = 0
    for r in range(2, ws.max_row + 1):
        title_raw = ws.cell(r, 5).value
        if not title_raw or _row_to_none([ws.cell(r, c).value for c in range(1, 13)]):
            continue
        if str(title_raw).strip().upper() in ('EN EMISIÓN', 'NOMBRE', 'EN EMISIÓ'):
            continue

        title, year = _parse_title_year(_clean(title_raw))
        if not title:
            continue

        country, network = _parse_country_network(_clean(ws.cell(r, 8).value))
        genres   = _clean_list(_clean(ws.cell(r, 7).value))
        rating   = _get_rating(ws.cell(r, 3))
        score_b  = ws.cell(r, 2).value
        score    = float(score_b) if score_b and str(score_b).replace('.','').isdigit() else None
        synopsis = _clean(ws.cell(r, 10).value)
        notes    = _clean(ws.cell(r, 11).value)
        platform = _clean(ws.cell(r, 12).value)
        duration = _clean(ws.cell(r, 6).value)
        cast     = _clean(ws.cell(r, 9).value)

        media_data = {
            'type': 'DORAMA', 'title': title, 'title_original': None,
            'year': year, 'genres': genres or None, 'synopsis': synopsis,
            'cover_url': None, 'duration': duration, 'country': country,
            'network': network, 'cast_text': cast, 'external_score': score,
            'platform': platform,
        }
        entry_data = {
            'status': status, 'progress': None, 'score': None,
            'rating_label': rating, 'notes': notes, 'platform': platform,
        }

        if not dry_run:
            mid = upsert_media(cur, media_data)
            upsert_entry(cur, user_id, mid, entry_data)
        count += 1
        if count % 50 == 0:
            log.info(f"    {count} doramas procesados...")

    return count


def parse_peliculas(ws, user_id: int, status: str, cur, dry_run: bool) -> int:
    """Peliculas — B=imdb  C=mine_color  D=title  E=year  F=duration  G=genres  H=synopsis  I=platform"""
    count = 0
    for r in range(2, ws.max_row + 1):
        title_raw = ws.cell(r, 4).value
        if not title_raw or _row_to_none([ws.cell(r, c).value for c in range(1, 11)]):
            continue

        title    = _clean(title_raw)
        if not title:
            continue
        year_raw = ws.cell(r, 5).value
        year     = int(float(str(year_raw))) if year_raw else None
        genres   = _clean_list(_clean(ws.cell(r, 7).value), sep=', ')
        rating   = _get_rating(ws.cell(r, 3))
        score_b  = ws.cell(r, 2).value
        score    = float(score_b) if score_b and str(score_b).replace('.','').isdigit() else None
        synopsis = _clean(ws.cell(r, 8).value)
        platform = _clean(ws.cell(r, 9).value)
        duration = _clean(ws.cell(r, 6).value)

        media_data = {
            'type': 'MOVIE', 'title': title, 'title_original': None,
            'year': year, 'genres': genres or None, 'synopsis': synopsis,
            'cover_url': None, 'duration': duration, 'country': None,
            'network': None, 'cast_text': None, 'external_score': score,
            'platform': platform,
        }
        entry_data = {
            'status': status, 'progress': None, 'score': None,
            'rating_label': rating, 'notes': None, 'platform': platform,
        }

        if not dry_run:
            mid = upsert_media(cur, media_data)
            upsert_entry(cur, user_id, mid, entry_data)
        count += 1

    return count


def parse_series(ws, user_id: int, status: str, cur, dry_run: bool) -> int:
    """Series — B=imdb  C=mine_color  D=title  E=estado_obra  F=t_total  G=viendo  H=t_vistas  I=dur  J=genres  K=synopsis  L=platform"""
    count = 0
    for r in range(2, ws.max_row + 1):
        title_raw = ws.cell(r, 4).value
        if not title_raw or _row_to_none([ws.cell(r, c).value for c in range(1, 13)]):
            continue

        title    = _clean(title_raw)
        if not title:
            continue

        genres   = _clean_list(_clean(ws.cell(r, 10).value), sep=', ')
        rating   = _get_rating(ws.cell(r, 3))
        score_b  = ws.cell(r, 2).value
        score    = float(score_b) if score_b and str(score_b).replace('.','').isdigit() else None
        synopsis = _clean(ws.cell(r, 11).value)
        platform = _clean(ws.cell(r, 12).value)
        duration = _clean(ws.cell(r, 9).value)
        progress_raw = _clean(ws.cell(r, 7).value)  # "Viendo", "Al día", "Finalizada"
        t_vistas = _clean(ws.cell(r, 8).value)
        progress = None
        if t_vistas:
            progress = f"T{int(float(str(t_vistas)))} vistas" if str(t_vistas).replace('.','').isdigit() else str(t_vistas)

        media_data = {
            'type': 'SERIES', 'title': title, 'title_original': None,
            'year': None, 'genres': genres or None, 'synopsis': synopsis,
            'cover_url': None, 'duration': duration, 'country': None,
            'network': None, 'cast_text': None, 'external_score': score,
            'platform': platform,
        }
        entry_data = {
            'status': status, 'progress': progress, 'score': None,
            'rating_label': rating, 'notes': None, 'platform': platform,
        }

        if not dry_run:
            mid = upsert_media(cur, media_data)
            upsert_entry(cur, user_id, mid, entry_data)
        count += 1

    return count


def parse_mangas(ws, user_id: int, status: str, col_offset: int, cur, dry_run: bool) -> int:
    """Mangas — WatchList: D=name  E=caps  F=genres  G=type  H=synopsis  I=link
       Mangas vistas:       C=name  D=caps  E=genres  F=type  G=synopsis
       col_offset: 0 para vistas (C=col3), 1 para watchlist (D=col4)
    """
    base = 3 + col_offset  # col C=3 (vistas), col D=4 (watchlist)
    count = 0
    for r in range(2, ws.max_row + 1):
        title_raw = ws.cell(r, base).value
        if not title_raw or _row_to_none([ws.cell(r, c).value for c in range(1, 10)]):
            continue
        if str(title_raw).strip().upper() in ('EN EMISIÓN', 'NOMBRE'):
            continue

        title, year = _parse_title_year(_clean(title_raw))
        if not title:
            continue

        caps_raw = ws.cell(r, base + 1).value
        caps     = _clean(caps_raw)
        genres   = _clean_list(_clean(ws.cell(r, base + 2).value))
        fmt_raw  = _clean(ws.cell(r, base + 3).value) or 'Manga'
        synopsis = _clean(ws.cell(r, base + 4).value)

        # Determine format from type text
        fmt_lower = fmt_raw.lower()
        if 'manhwa' in fmt_lower:
            media_type = 'MANHWA'
        elif 'manhua' in fmt_lower:
            media_type = 'MANHUA'
        elif 'webtoon' in fmt_lower:
            media_type = 'WEBTOON'
        else:
            media_type = 'MANGA'

        rating   = _get_rating(ws.cell(r, 3))
        progress = caps if caps and status != 'completed' else None

        media_data = {
            'type': media_type, 'title': title, 'title_original': None,
            'year': year, 'genres': genres or None, 'synopsis': synopsis,
            'cover_url': None, 'duration': None, 'country': None,
            'network': None, 'cast_text': None, 'external_score': None,
            'platform': None,
        }
        entry_data = {
            'status': status, 'progress': progress, 'score': None,
            'rating_label': rating, 'notes': None, 'platform': None,
        }

        if not dry_run:
            mid = upsert_media(cur, media_data)
            upsert_entry(cur, user_id, mid, entry_data)
        count += 1

    return count


def parse_animes(ws, user_id: int, status: str, cur, dry_run: bool) -> int:
    """Animes — B=mine_color  C=title  D=estado  E=t_total  F=viendo  G=pais  H=genres  I=synopsis  J=platform"""
    count = 0
    for r in range(2, ws.max_row + 1):
        title_raw = ws.cell(r, 3).value
        if not title_raw or _row_to_none([ws.cell(r, c).value for c in range(1, 11)]):
            continue

        title    = _clean(title_raw)
        if not title:
            continue

        genres   = _clean_list(_clean(ws.cell(r, 8).value), sep=', ')
        rating   = _get_rating(ws.cell(r, 2))
        synopsis = _clean(ws.cell(r, 9).value)
        platform = _clean(ws.cell(r, 10).value)
        country  = _clean(ws.cell(r, 7).value)
        progress_raw = _clean(ws.cell(r, 6).value)

        media_data = {
            'type': 'ANIME', 'title': title, 'title_original': None,
            'year': None, 'genres': genres or None, 'synopsis': synopsis,
            'cover_url': None, 'duration': None, 'country': country,
            'network': None, 'cast_text': None, 'external_score': None,
            'platform': platform,
        }
        entry_data = {
            'status': status, 'progress': progress_raw, 'score': None,
            'rating_label': rating, 'notes': None, 'platform': platform,
        }

        if not dry_run:
            mid = upsert_media(cur, media_data)
            upsert_entry(cur, user_id, mid, entry_data)
        count += 1

    return count


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--excel', default='/opt/docker/zonatmo/Doramas, manga y anime.xlsx')
    parser.add_argument('--user', default='laura')
    parser.add_argument('--dry-run', action='store_true')
    args = parser.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.exists():
        log.error(f"Excel no encontrado: {excel_path}")
        sys.exit(1)

    if args.dry_run:
        log.info("=== DRY RUN — no se escribirá en la DB ===")

    log.info(f"Cargando Excel: {excel_path}")
    wb = openpyxl.load_workbook(str(excel_path))

    conn = None if args.dry_run else _conn()
    cur  = None if args.dry_run else conn.cursor()

    try:
        user_id = 0
        if not args.dry_run:
            user_id = get_or_create_user(cur, args.user)
            log.info(f"Usuario: {args.user} (id={user_id})")

        stats = {}

        # Doramas vistas
        log.info("Importando Doramas (vistas)...")
        n = parse_doramas(wb['Doramas'], user_id, 'completed', cur, args.dry_run)
        stats['Doramas vistas'] = n
        if not args.dry_run: conn.commit()

        # Doramas WatchList
        log.info("Importando Doramas WatchList...")
        n = parse_doramas(wb['Doramas WatchList'], user_id, 'plan_to_watch', cur, args.dry_run)
        stats['Doramas watchlist'] = n
        if not args.dry_run: conn.commit()

        # Películas
        log.info("Importando Películas...")
        n = parse_peliculas(wb['Peliculas'], user_id, 'completed', cur, args.dry_run)
        stats['Películas vistas'] = n
        n = parse_peliculas(wb['Peliculas WatchList'], user_id, 'plan_to_watch', cur, args.dry_run)
        stats['Películas watchlist'] = n
        if not args.dry_run: conn.commit()

        # Series
        log.info("Importando Series...")
        n = parse_series(wb['Series'], user_id, 'completed', cur, args.dry_run)
        stats['Series vistas'] = n
        n = parse_series(wb['Series WatchList'], user_id, 'plan_to_watch', cur, args.dry_run)
        stats['Series watchlist'] = n
        if not args.dry_run: conn.commit()

        # Mangas
        log.info("Importando Mangas...")
        n = parse_mangas(wb['Mangas'], user_id, 'completed', 0, cur, args.dry_run)
        stats['Mangas leídos'] = n
        n = parse_mangas(wb['Mangas WatchList'], user_id, 'watching', 1, cur, args.dry_run)
        stats['Mangas en progreso'] = n
        if not args.dry_run: conn.commit()

        # Animes
        log.info("Importando Animes...")
        n = parse_animes(wb['Animes'], user_id, 'completed', cur, args.dry_run)
        stats['Animes vistos'] = n
        n = parse_animes(wb['Animes WatchList'], user_id, 'plan_to_watch', cur, args.dry_run)
        stats['Animes watchlist'] = n
        if not args.dry_run: conn.commit()

    except Exception as e:
        if conn:
            conn.rollback()
        log.error(f"Error: {e}")
        raise
    finally:
        if cur:  cur.close()
        if conn: conn.close()

    log.info("\n=== RESUMEN ===")
    total = 0
    for k, v in stats.items():
        log.info(f"  {k}: {v}")
        total += v
    log.info(f"  TOTAL: {total} entradas")
    if args.dry_run:
        log.info("  [DRY RUN — nada importado]")
    else:
        log.info("  ✓ Importación completada")


if __name__ == '__main__':
    main()
